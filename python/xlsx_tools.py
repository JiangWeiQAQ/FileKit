import csv
from pathlib import Path
from typing import Any, Dict, List


def require_openpyxl() -> Any:
    try:
        import openpyxl
        return openpyxl
    except ImportError as error:
        raise RuntimeError("缺少 Python 包 openpyxl，请安装后重试。") from error


def load(path: str) -> Any:
    openpyxl = require_openpyxl()
    return openpyxl.load_workbook(path, read_only=False, data_only=False, keep_vba=path.lower().endswith(".xlsm"))


def info(path: str) -> Dict[str, Any]:
    workbook = load(path)
    sheets = [{"name": sheet.title, "maxRow": sheet.max_row, "maxColumn": sheet.max_column, "dimension": sheet.calculate_dimension()} for sheet in workbook.worksheets]
    return {"sheetCount": len(sheets), "sheets": sheets}


def export_csv(path: str, sheet_name: str | None, all_sheets: bool, create_output: Any) -> Dict[str, Any]:
    workbook = load(path)
    sheets = workbook.worksheets if all_sheets else [workbook[sheet_name]] if sheet_name in workbook.sheetnames else []
    if not sheets:
        return {"success": False, "message": f"找不到工作表：{sheet_name}。", "files": [], "data": {"sheets": workbook.sheetnames}}
    files: List[str] = []
    for sheet in sheets:
        output = create_output(Path(path).stem, f"{sheet.title}_export", "csv")
        with open(output, "w", encoding="utf-8-sig", newline="") as target:
            writer = csv.writer(target)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        files.append(output)
    return {"success": True, "message": f"已导出 {len(files)} 个 CSV 文件。", "files": files, "data": {"sheetNames": [sheet.title for sheet in sheets]}}


def clean_blank_rows_columns(path: str, create_output: Any) -> Dict[str, Any]:
    workbook = load(path)
    removed_rows = 0
    removed_columns = 0
    for sheet in workbook.worksheets:
        for row_index in range(sheet.max_row, 0, -1):
            if all(cell.value is None for cell in sheet[row_index]):
                sheet.delete_rows(row_index)
                removed_rows += 1
        for column_index in range(sheet.max_column, 0, -1):
            if all(sheet.cell(row=row_index, column=column_index).value is None for row_index in range(1, sheet.max_row + 1)):
                sheet.delete_cols(column_index)
                removed_columns += 1
    output = create_output(Path(path).stem, "cleaned", "xlsx")
    workbook.save(output)
    return {"success": True, "message": f"已删除 {removed_rows} 个完全空白行和 {removed_columns} 个完全空白列；原文件未修改。", "files": [output], "data": {"removedRows": removed_rows, "removedColumns": removed_columns}}
